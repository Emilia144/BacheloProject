"""
transform.py – Pipeline: comp1.xlsx + results.xlsx → comparison_bcg.json

Reads raw survey data for a company subset (comp1.xlsx) and the full dataset
(results.xlsx), computes all metrics used by vergleich.html, and writes a
structured JSON file that JS diagram functions can consume.

Usage:
    python data/transform.py                         # default files
    python data/transform.py --company data/comp1.xlsx --overall data/results.xlsx
"""

from __future__ import annotations

import argparse
import json
import math
import os
import sys
from collections import Counter
from pathlib import Path

import openpyxl


# ─── Column mapping (1-indexed, matching the Excel layout) ───────────────────

COL = {
    "id":              1,   # A
    "bereich":         2,   # B
    "groesse":         3,   # C
    "branche":         4,   # D
    "frequenz":        5,   # E  – Nutzungshäufigkeit
    "zwecke":          6,   # F  – Nutzungszwecke (multi-select, ";"-separated)
    "tools":           7,   # G  – GenAI Tools (multi-select, ";"-separated)
    "kompetenz":       8,   # H  – Kompetenzstufe
    "prod_heute":      9,   # I  – Produktivität heute
    "prod_zukunft":   10,   # J  – Produktivität in 1-3 Jahren
    "einordnung":     11,   # K  – Einordnung KI-Ergebnisse
    "warum_nicht":    12,   # L  – Warum keine GenAI
    "verbesserungen": 13,   # M  – Verbesserungswünsche
    "datenschutz":    14,   # N  – Datenschutz
    "strategisch":    15,   # O  – Strategische Wirkung
    "phase":          16,   # P  – Implementierungsphase
    "herausforderung":17,   # Q  – Herausforderungen
    "modellstrategie":18,   # R  – Modellstrategie
}


# ─── Helpers ─────────────────────────────────────────────────────────────────

def read_rows(path: str | Path, sheet: str | None = None) -> list[dict]:
    """Read an xlsx file and return a list of dicts keyed by column index."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    if sheet:
        ws = wb[sheet]
    else:
        # Try "Sheet1" first, fall back to active sheet
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for col_idx in range(1, ws.max_column + 1):
            row[col_idx] = ws.cell(r, col_idx).value
        rows.append(row)
    return rows


def pct(count: int, total: int, decimals: int = 1) -> float:
    """Safe percentage with rounding."""
    if total == 0:
        return 0.0
    return round(count / total * 100, decimals)


def delta(company_pct: float, overall_pct: float, decimals: int = 1) -> float:
    """Delta in percentage points."""
    return round(company_pct - overall_pct, decimals)


def parse_multi(value: str | None) -> list[str]:
    """Split a semicolon-separated multi-select field into items."""
    if not value:
        return []
    return [item.strip() for item in value.split(";") if item.strip()]


# ─── Analysis functions ──────────────────────────────────────────────────────

def analyse_usage_frequency(rows: list[dict]) -> dict:
    """Compute usage frequency distribution (Täglich, Wöchentlich, Selten, Gar nicht)."""
    n = len(rows)
    categories = ["Täglich", "Wöchentlich", "Selten", "Gar nicht"]
    counter = Counter()
    for row in rows:
        val = row[COL["frequenz"]]
        if val:
            # Normalize encoding
            normalized = val.replace("ä", "ä").replace("\u00e4", "ä")
            counter[normalized] += 1

    result = {}
    for cat in categories:
        count = counter.get(cat, 0)
        result[cat] = {"count": count, "percentage": pct(count, n)}
    return result


def analyse_adoption(rows: list[dict]) -> dict:
    """Compute adoption rate (% that use GenAI at all, i.e. not 'Gar nicht')."""
    n = len(rows)
    non_users = sum(1 for r in rows if r[COL["frequenz"]] and "Gar nicht" in r[COL["frequenz"]])
    users = n - non_users
    return {
        "total": n,
        "users": users,
        "non_users": non_users,
        "adoption_rate": pct(users, n),
    }


def analyse_tools(rows: list[dict]) -> dict:
    """Compute tool usage from multi-select field (% of users who selected each tool)."""
    # Only count users (those who have tools listed)
    user_rows = [r for r in rows if r[COL["tools"]]]
    n_users = len(user_rows)

    tool_map = {
        "OpenAI ChatGPT": "ChatGPT",
        "Microsoft Copilot": "Microsoft Copilot",
        "Google Gemini": "Google Gemini",
        "Anthrophic Claude": "Anthropic Claude",
        "Lokale OpenSource Modelle": "Local/Open Source",
        "Andere cloudbasierte Modelle": "Andere Cloud-Tools",
    }

    counter = Counter()
    for row in user_rows:
        items = parse_multi(row[COL["tools"]])
        seen = set()
        for item in items:
            for key, label in tool_map.items():
                if key in item and label not in seen:
                    counter[label] += 1
                    seen.add(label)

    result = {}
    for label in tool_map.values():
        count = counter.get(label, 0)
        result[label] = {"count": count, "percentage": pct(count, n_users)}
    return result


def analyse_productivity(rows: list[dict], column_key: str) -> dict:
    """Compute productivity distribution for a given column (today or future).

    Only counts users (rows where the column has a value).
    """
    col = COL[column_key]
    user_rows = [r for r in rows if r[col]]
    n = len(user_rows)

    mapping = {
        "Deutliche Verbesserung": "significant",
        "Leichte Verbesserung": "slight",
        "Keine Änderung": "none",
        "Keine Einschätzung möglich": "none",
        "Negative Auswirkung": "negative",
    }

    counter = Counter()
    for row in user_rows:
        val = row[col]
        # Try each mapping key
        matched = False
        for text, key in mapping.items():
            if text in val:
                counter[key] += 1
                matched = True
                break
        if not matched:
            counter["none"] += 1

    result = {
        "n": n,
        "significant": pct(counter.get("significant", 0), n),
        "slight": pct(counter.get("slight", 0), n),
        "none": pct(counter.get("none", 0), n),
        "negative": pct(counter.get("negative", 0), n),
    }
    result["total_positive"] = round(result["significant"] + result["slight"], 1)
    result["none_or_negative"] = round(result["none"] + result["negative"], 1)
    return result


def analyse_competence(rows: list[dict]) -> dict:
    """Compute competence level distribution across 6 tiers."""
    col = COL["kompetenz"]
    user_rows = [r for r in rows if r[col]]
    n = len(user_rows)

    # Map answer text fragments to tier labels
    tiers = [
        ("ausprobiert", "Ausprobiert"),
        ("einfache Prompts", "Einfache Prompts"),
        ("verschiedene Modelle", "Modelle & Prompting"),
        ("erweiterte Funktionen", "Erweiterte Funktionen"),
        ("eigene GenAI-basierte", "Eigene Lösungen"),
        ("eigene Modelle", "Eigene Modelle"),
    ]

    counter = Counter()
    for row in user_rows:
        val = row[col].lower() if row[col] else ""
        for fragment, label in tiers:
            if fragment.lower() in val:
                counter[label] += 1
                break

    result = {}
    for _, label in tiers:
        count = counter.get(label, 0)
        result[label] = {"count": count, "percentage": pct(count, n)}

    # Aggregates: low (tiers 1-2) vs mid-high (tiers 3-6)
    low = sum(counter.get(label, 0) for _, label in tiers[:2])
    mid_high = sum(counter.get(label, 0) for _, label in tiers[2:])
    result["_low_pct"] = pct(low, n)
    result["_mid_high_pct"] = pct(mid_high, n)
    result["_n"] = n
    return result


def analyse_implementation(rows: list[dict]) -> dict:
    """Compute implementation phase distribution (4 phases)."""
    col = COL["phase"]
    resp_rows = [r for r in rows if r[col]]
    n = len(resp_rows)

    phases = [
        ("Konzeption", "phase1", "Konzeption & Entwicklung"),
        ("Prototyp", "phase2", "Prototyp vorhanden"),
        ("Produktivversion", "phase3", "Erste Produktivversion"),
        ("Rollout", "phase4", "Unternehmensweiter Rollout"),
    ]

    counter = Counter()
    for row in resp_rows:
        val = row[col]
        for fragment, key, _ in phases:
            if fragment in val:
                counter[key] += 1
                break

    result = {}
    for fragment, key, label in phases:
        count = counter.get(key, 0)
        result[key] = {"count": count, "percentage": pct(count, n), "label": label}

    p3 = result["phase3"]["percentage"]
    p4 = result["phase4"]["percentage"]
    result["phase34_combined"] = round(p3 + p4, 1)
    result["_n"] = n
    return result


def analyse_challenges(rows: list[dict]) -> list[dict]:
    """Compute challenge ranking (single-select)."""
    col = COL["herausforderung"]
    resp_rows = [r for r in rows if r[col]]
    n = len(resp_rows)

    # Short labels for known challenges
    label_map = {
        "Fehlende Datenqualität": "Fehlende Datenqualität und -verfügbarkeit",
        "Unklarer Business Case": "Unklarer Business Case / ROI-Nachweis",
        "Integration in bestehende": "Integration in bestehende IT-Systeme",
        "Mangel an qualifiziertem": "Mangel an qualifiziertem Personal und Know-how",
        "Unklare Governance": "Unklare Governance und regulatorische Unsicherheit",
    }

    counter = Counter()
    for row in resp_rows:
        val = row[col]
        for fragment, label in label_map.items():
            if fragment in val:
                counter[label] += 1
                break

    ranked = []
    for rank, (challenge, count) in enumerate(counter.most_common(), start=1):
        ranked.append({
            "rank": rank,
            "challenge": challenge,
            "count": count,
            "percentage": pct(count, n),
        })
    return ranked


def analyse_privacy(rows: list[dict]) -> dict:
    """Compute data privacy awareness distribution."""
    col = COL["datenschutz"]
    resp_rows = [r for r in rows if r[col]]
    n = len(resp_rows)

    # Order matters: more specific fragments must come first to avoid
    # "nicht immer konsequent" matching "konsequent" prematurely.
    categories = [
        ("nicht immer konsequent", "Richtlinien bekannt, nicht immer umgesetzt"),
        ("keine konkreten Regeln", "Keine konkreten Regeln bekannt"),
        ("besten Gewissens", "Wichtigste Regelungen bekannt, best-effort"),
        ("konsequent", "Regelungen bekannt & konsequent umgesetzt"),
    ]

    counter = Counter()
    for row in resp_rows:
        val = row[col]
        for fragment, label in categories:
            if fragment in val:
                counter[label] += 1
                break

    # Output in display order (konsequent first, keine Regeln last)
    display_order = [
        "Regelungen bekannt & konsequent umgesetzt",
        "Wichtigste Regelungen bekannt, best-effort",
        "Richtlinien bekannt, nicht immer umgesetzt",
        "Keine konkreten Regeln bekannt",
    ]
    result = {}
    for label in display_order:
        count = counter.get(label, 0)
        result[label] = {"count": count, "percentage": pct(count, n)}

    # Aggregate: top-2 = "compliant"
    top2 = sum(counter.get(label, 0) for label in display_order[:2])
    result["_compliant_pct"] = pct(top2, n)
    result["_n"] = n
    return result


def analyse_strategic_impact(rows: list[dict]) -> dict:
    """Compute strategic impact (Effizienzhebel vs Wachstumstreiber)."""
    col = COL["strategisch"]
    resp_rows = [r for r in rows if r[col]]
    n = len(resp_rows)

    categories = [
        ("Effizienzhebel", "Effizienzhebel"),
        ("Gleichermaßen beides", "Gleichermaßen beides"),
        ("Wachstumstreiber", "Wachstumstreiber"),
        ("weder noch", "Bisher weder noch erkennbar"),
    ]

    counter = Counter()
    for row in resp_rows:
        val = row[col]
        for fragment, label in categories:
            if fragment.lower() in val.lower():
                counter[label] += 1
                break

    result = {}
    for _, label in categories:
        count = counter.get(label, 0)
        result[label] = {"count": count, "percentage": pct(count, n)}
    result["_n"] = n
    return result


def analyse_nutzungszwecke(rows: list[dict]) -> dict:
    """Compute usage purposes from multi-select field (% of users)."""
    col = COL["zwecke"]
    user_rows = [r for r in rows if r[col]]
    n = len(user_rows)

    counter = Counter()
    for row in user_rows:
        items = parse_multi(row[col])
        seen = set()
        for item in items:
            # Normalize to a short label
            label = item.split("(")[0].strip().rstrip(";")
            if label and label not in seen:
                counter[label] += 1
                seen.add(label)

    result = {}
    for label, count in counter.most_common():
        result[label] = {"count": count, "percentage": pct(count, n)}
    result["_n"] = n
    return result


# ─── Comparison builder ──────────────────────────────────────────────────────

def build_comparison(company_rows: list[dict], overall_rows: list[dict]) -> dict:
    """Run all analyses for both groups and produce a comparison dict."""

    n_company = len(company_rows)
    n_overall = len(overall_rows)

    # ── Run all analyses ──
    c_adoption = analyse_adoption(company_rows)
    o_adoption = analyse_adoption(overall_rows)

    c_freq = analyse_usage_frequency(company_rows)
    o_freq = analyse_usage_frequency(overall_rows)

    c_tools = analyse_tools(company_rows)
    o_tools = analyse_tools(overall_rows)

    c_prod_today = analyse_productivity(company_rows, "prod_heute")
    o_prod_today = analyse_productivity(overall_rows, "prod_heute")

    c_prod_future = analyse_productivity(company_rows, "prod_zukunft")
    o_prod_future = analyse_productivity(overall_rows, "prod_zukunft")

    c_competence = analyse_competence(company_rows)
    o_competence = analyse_competence(overall_rows)

    c_impl = analyse_implementation(company_rows)
    o_impl = analyse_implementation(overall_rows)

    c_challenges = analyse_challenges(company_rows)
    o_challenges = analyse_challenges(overall_rows)

    c_privacy = analyse_privacy(company_rows)
    o_privacy = analyse_privacy(overall_rows)

    c_strategic = analyse_strategic_impact(company_rows)
    o_strategic = analyse_strategic_impact(overall_rows)

    c_zwecke = analyse_nutzungszwecke(company_rows)
    o_zwecke = analyse_nutzungszwecke(overall_rows)

    # ── Build output ──
    output = {
        "sample_sizes": {
            "company": n_company,
            "overall": n_overall,
            "company_users": c_adoption["users"],
            "overall_users": o_adoption["users"],
        },
        "adoption": {
            "company": c_adoption["adoption_rate"],
            "overall": o_adoption["adoption_rate"],
            "delta": delta(c_adoption["adoption_rate"], o_adoption["adoption_rate"]),
        },
        "usage_frequency": _merge_categories(c_freq, o_freq),
        "tool_usage": _merge_categories(c_tools, o_tools),
        "nutzungszwecke": _merge_categories(c_zwecke, o_zwecke),
        "productivity_today": {
            "company": _strip_internal(c_prod_today),
            "overall": _strip_internal(o_prod_today),
        },
        "productivity_future": {
            "company": _strip_internal(c_prod_future),
            "overall": _strip_internal(o_prod_future),
        },
        "competence": _merge_categories(c_competence, o_competence),
        "implementation_phases": _merge_impl(c_impl, o_impl),
        "challenges": {
            "company": c_challenges,
            "overall": o_challenges,
        },
        "privacy": _merge_categories(c_privacy, o_privacy),
        "strategic_impact": _merge_categories(c_strategic, o_strategic),
        "kpi_cards": _build_kpi_cards(
            c_adoption, o_adoption,
            c_freq, o_freq,
            c_zwecke, o_zwecke,
            c_prod_today, o_prod_today,
            c_impl, o_impl,
            c_strategic, o_strategic,
        ),
    }
    return output


def _merge_categories(company: dict, overall: dict) -> dict:
    """Merge company/overall dicts by category, adding deltas."""
    result = {}
    for key in company:
        if key.startswith("_"):
            continue
        c = company[key]
        o = overall.get(key, {})
        if isinstance(c, dict) and "percentage" in c:
            c_pct = c["percentage"]
            o_pct = o["percentage"] if isinstance(o, dict) else 0.0
            result[key] = {
                "company": c_pct,
                "overall": o_pct,
                "delta": delta(c_pct, o_pct),
            }
        else:
            result[key] = {"company": c, "overall": overall.get(key)}
    return result


def _merge_impl(company: dict, overall: dict) -> dict:
    """Merge implementation phase data."""
    result = {}
    for key in ["phase1", "phase2", "phase3", "phase4"]:
        c = company[key]
        o = overall[key]
        result[key] = {
            "label": c["label"],
            "company": c["percentage"],
            "overall": o["percentage"],
            "delta": delta(c["percentage"], o["percentage"]),
        }
    result["phase34_combined"] = {
        "company": company["phase34_combined"],
        "overall": overall["phase34_combined"],
        "delta": delta(company["phase34_combined"], overall["phase34_combined"]),
    }
    return result


def _strip_internal(d: dict) -> dict:
    """Remove internal keys starting with _."""
    return {k: v for k, v in d.items() if not k.startswith("_")}


def _build_kpi_cards(
    c_adopt, o_adopt, c_freq, o_freq, c_zwecke, o_zwecke,
    c_prod, o_prod, c_impl, o_impl, c_strat, o_strat,
) -> list[dict]:
    """Build the 6 KPI card entries shown in the 'Auf einen Blick' section."""

    # Find top Nutzungszweck
    top_zweck = max(
        ((k, v) for k, v in c_zwecke.items() if not k.startswith("_")),
        key=lambda x: x[1]["percentage"],
        default=("", {"percentage": 0}),
    )
    top_zweck_label = top_zweck[0]
    c_zweck_pct = top_zweck[1]["percentage"]
    o_zweck_pct = o_zwecke.get(top_zweck_label, {}).get("percentage", 0.0) if isinstance(o_zwecke.get(top_zweck_label), dict) else 0.0

    # Strategic: total positive = Effizienz + Beides + Wachstum
    strat_keys = ["Effizienzhebel", "Gleichermaßen beides", "Wachstumstreiber"]
    c_strat_pos = sum(c_strat.get(k, {}).get("percentage", 0) for k in strat_keys if not k.startswith("_"))
    o_strat_pos = sum(o_strat.get(k, {}).get("percentage", 0) for k in strat_keys if not k.startswith("_"))

    cards = [
        {
            "id": "adoption",
            "title": "Nutzen GenAI im Arbeitskontext",
            "company": c_adopt["adoption_rate"],
            "overall": o_adopt["adoption_rate"],
            "delta": delta(c_adopt["adoption_rate"], o_adopt["adoption_rate"]),
        },
        {
            "id": "daily_usage",
            "title": "Nutzen GenAI täglich",
            "company": c_freq.get("Täglich", {}).get("percentage", 0),
            "overall": o_freq.get("Täglich", {}).get("percentage", 0),
            "delta": delta(
                c_freq.get("Täglich", {}).get("percentage", 0),
                o_freq.get("Täglich", {}).get("percentage", 0),
            ),
        },
        {
            "id": "top_purpose",
            "title": f"Nutzen GenAI für {top_zweck_label}",
            "company": c_zweck_pct,
            "overall": o_zweck_pct,
            "delta": delta(c_zweck_pct, o_zweck_pct),
        },
        {
            "id": "productivity",
            "title": "sehen eine Produktivitätssteigerung",
            "company": c_prod["total_positive"],
            "overall": o_prod["total_positive"],
            "delta": delta(c_prod["total_positive"], o_prod["total_positive"]),
        },
        {
            "id": "productive_use",
            "title": "Im Produktiveinsatz (Phase 3+4)",
            "company": c_impl["phase34_combined"],
            "overall": o_impl["phase34_combined"],
            "delta": delta(c_impl["phase34_combined"], o_impl["phase34_combined"]),
        },
        {
            "id": "strategic_positive",
            "title": "sehen GenAI als Effizienz- oder Wachstumshebel",
            "company": round(c_strat_pos, 1),
            "overall": round(o_strat_pos, 1),
            "delta": delta(c_strat_pos, o_strat_pos),
        },
    ]
    return cards


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Transform survey Excel data into comparison JSON."
    )
    script_dir = Path(__file__).resolve().parent
    parser.add_argument(
        "--company",
        default=str(script_dir / "comp1.xlsx"),
        help="Path to the company subset Excel file (default: data/comp1.xlsx)",
    )
    parser.add_argument(
        "--overall",
        default=str(script_dir / "results.xlsx"),
        help="Path to the full dataset Excel file (default: data/results.xlsx)",
    )
    parser.add_argument(
        "--output",
        default=str(script_dir / "comparison_bcg.json"),
        help="Output JSON path (default: data/comparison_bcg.json)",
    )
    parser.add_argument(
        "--pretty",
        action="store_true",
        default=True,
        help="Pretty-print JSON (default: True)",
    )
    args = parser.parse_args()

    print(f"Reading company data: {args.company}")
    company_rows = read_rows(args.company)
    print(f"  {len(company_rows)} rows")

    print(f"Reading overall data: {args.overall}")
    overall_rows = read_rows(args.overall)
    print(f"  {len(overall_rows)} rows")

    print("Running analyses...")
    result = build_comparison(company_rows, overall_rows)

    indent = 2 if args.pretty else None
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=indent)

    print(f"Written: {args.output}")


if __name__ == "__main__":
    main()
